import tkinter as tk
from tkinter import messagebox

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from numpy import array

import openpyxl
from openpyxl.chart import BarChart, Reference

# Đọc file và lấy dữ liệu
df = pd.read_csv('diemPython.csv', index_col=0, header=0)
in_data = array(df.iloc[:, :])

# Lấy ra tổng sinh viên
tongsv = np.sum(in_data[:, 1])

# Lấy ra tất cả các điểm
diemA = in_data[:, 3]
diemB_plus = in_data[:, 4]
diemB = in_data[:, 5]
diemC_plus = in_data[:, 6]
diemC = in_data[:, 7]
diemD_plus = in_data[:, 8]
diemD = in_data[:, 9]
diemF = in_data[:, 10]


# Hàm trả về % số sinh viên đi thi đạt điểm F
def percentFailure(array, tongSV):
    sum = np.sum(array)
    average = (sum / tongSV) * 100
    return average


# Hàm in ra lớp có nhiều sinh viên giỏi nhất
def bestLevel(array1, array2):
    arrayGood = np.add(array1, array2)
    max = np.max(arrayGood)
    i, = np.where(arrayGood == max)
    print('Lớp có nhiều học giỏi nhất là {0} có {1} sinh viên'.format(in_data[i, 0], max))


# Hàm in ra điểm cao nhất của từng lớp
def maxScoreOfClass(a):
    caclop = in_data[a, 2:11]
    print(caclop)
    maxb = caclop.max()
    maxbb = maxb
    if maxbb == in_data[a, 2]:
        aa = "Loai A+"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(aa, maxb))
    elif maxbb == in_data[a, 3]:
        b = "Loai A"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(b, maxb))
    elif maxbb == in_data[a, 4]:
        c = "Loai B+"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(c, maxb))
    elif maxbb == in_data[a, 5]:
        d = "Loai B"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(d, maxb))
    elif maxbb == in_data[a, 6]:
        e = "Loai C+"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(e, maxb))
    elif maxbb == in_data[a, 7]:
        f = "Loai C"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(f, maxb))
    elif maxbb == in_data[a, 8]:
        g = "Loai D+"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(g, maxb))
    elif maxbb == in_data[a, 9]:
        h = "Loai D"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(h, maxb))
    elif maxbb == in_data[a, 10]:
        j = "Loai F"
        print('lop co diem cao nhat la diem {0} va co {1} sinh vien dat duoc'.format(j, maxb))


# Hàm in ra điểm thấp nhất của từng lớp
def minScoreOfClass(a):
    caclop = in_data[a, 2:11]
    print(caclop)
    maxb = caclop.min()
    maxbb = maxb
    if maxbb == in_data[a, 2]:
        aa = "Loai A+"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(aa, maxb))
    elif maxbb == in_data[a, 3]:
        b = "Loai A"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(b, maxb))
    elif maxbb == in_data[a, 4]:
        c = "Loai B+"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(c, maxb))
    elif maxbb == in_data[a, 5]:
        d = "Loai B"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(d, maxb))
    elif maxbb == in_data[a, 6]:
        e = "Loai C+"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(e, maxb))
    elif maxbb == in_data[a, 7]:
        f = "Loai C"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(f, maxb))
    elif maxbb == in_data[a, 8]:
        g = "Loai D+"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(g, maxb))
    elif maxbb == in_data[a, 9]:
        h = "Loai D"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(h, maxb))
    elif maxbb == in_data[a, 10]:
        j = "Loai F"
        print('lop co diem thap nhat la diem {0} va co {1} sinh vien dat duoc'.format(j, maxb))


# Vẽ đồ thị so sánh số sinh viên điểm A mỗi lớp so với trung bình số sinh viên đạt điểm A
def TBC_A():
    diemA = in_data[:, 3]
    tongA = np.sum(diemA)
    soLop = len(in_data[:, 0])
    avg = tongA / soLop
    avgArr = np.zeros(1, soLop)
    for i in range(soLop):
        avgArr[i] = avg
    plt.plot(range(len(diemA)), diemA, 'r-', label="Diem A")
    plt.plot(range(len(diemA)), avgArr, '-', label="TB diem A")
    plt.xlabel('Lơp')
    plt.ylabel(' So sv dat diem A')
    plt.legend(loc='upper right')
    plt.show()

# cirleCompareScore(diemA, diemB_plus, diemB, diemC_plus, diemC, diemD_plus, diemD, diemF)
# print('Tong sv:',tongsv)
# maxa = diemA.max()
# i = np.where(diemA == maxa)
# print('lop co nhieu diem A la {0} co {1} sv dat diem A'.format(in_data[i,0],maxa))
# chonlop=int(input("nhap stt lop muon chon: "))
# diemcaonhatcuatunglop(chonlop)
# chonlopthapnhat=int(input("nhap stt lop muon chon: "))
# diemthapnhatcuatunglop(chonlopthapnhat)

# Hàm xử lý hiển thị thông tin dưới dạng text với từng lớp
def solve():
    try:
        num_equations = int(entry.get())
        if num_equations >= 1 and num_equations <= 9:
            students = in_data[num_equations - 1, 1]
            lb1 = tk.Label(window,
                           text="Số sinh viên của lớp " + in_data[num_equations - 1, 0] + " là: " + str(students))
            lb1.grid(row=3, column=0)
            good_student = in_data[num_equations - 1, 3] + in_data[num_equations - 1, 4]
            lb2 = tk.Label(window, text="Số sinh viên giỏi: " + str(good_student))
            lb2.grid(row=4, column=0)
            bad_student = in_data[num_equations - 1, 10]
            lb3 = tk.Label(window, text="Số sinh viên kém: " + str(bad_student))
            lb3.grid(row=5, column=0)
        else:
            messagebox.showerror("Error", "Lớp không tồn tại")
    except Exception as e:
        messagebox.showerror("Error", e)

# Hàm xử lý hiển thị thông tin dưới dạng biểu đồ với từng lớp
def solve1():
    try:
        plt.close()
        num_equations = int(entry.get())
        if num_equations >= 1 and num_equations <= 9:
            scoreA = in_data[num_equations - 1, 3]
            scoreB_plus = in_data[num_equations - 1, 4]
            scoreB = in_data[num_equations - 1, 5]
            scoreC_plus = in_data[num_equations - 1, 6]
            scoreC = in_data[num_equations - 1, 7]
            scoreD_plus = in_data[num_equations - 1, 8]
            scoreD = in_data[num_equations - 1, 9]
            scoreF = in_data[num_equations - 1, 10]
            labels = ['A', 'B+', 'B', 'C+', 'C', 'D+', 'D', 'F']
            sizes = [scoreA, scoreB_plus, scoreB, scoreC_plus,
                     scoreC, scoreD_plus, scoreD, scoreF]
            plt.pie(sizes, labels=labels,autopct='%1.1f%%', startangle=140)
            # Vẽ biểu đồ thanh
            plt.title('Biểu đồ đánh giá số điểm lớp')
            plt.show()
        else:
            messagebox.showerror("Error", "Lớp không tồn tại")
    except Exception as e:
        messagebox.showerror("Error", e)

# Hàm vẽ biểu đồ tròn so sánh tổng số điểm của các lớp
def solve2():
        plt.close()
        averageSumA = np.sum(diemA)
        averageSumB_plus = np.sum(diemB_plus)
        averageSumB = np.sum(diemB)
        averageSumC_plus = np.sum(diemC_plus)
        averageSumC = np.sum(diemC)
        averageSumD_plus = np.sum(diemD_plus)
        averageSumD = np.sum(diemD)
        averageSumF = np.sum(diemF)

        labels = ['A', 'B+', 'B', 'C+', 'C', 'D+', 'D', 'F']
        sizes = [averageSumA, averageSumB_plus, averageSumB, averageSumC_plus,
                 averageSumC, averageSumD_plus, averageSumD, averageSumF]
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.title('Biểu đồ đánh giá tất cả sinh viên đi thi môn học')
        plt.show()

# Lấy hàng điểm của lớp đầu tiên: print(in_data[0, 3:11])
# Tạo cửa sổ giao diện
window = tk.Tk()
window.geometry('1000x1000')
window.title("Phân tích kết quả kết thúc học phần môn học")

lb = tk.Label(window, text="Nhập vào số thứ tự lớp muốn phân tích:")
lb.grid(row=0, column=0)

entry = tk.Entry(window)
entry.grid(row=0, column=1, padx=10)

btn1 = tk.Button(window, text="Xem thông tin", command=solve)
btn1.grid(row=2, column=0, ipadx=5, ipady=5, pady=10)

btn2 = tk.Button(window, text="Xem biểu đồ", command=solve1)
btn2.grid(row=2, column=1, ipadx=5, ipady=5)

btn3 = tk.Button(window, text="Xem biểu đồ phân tích chung", command=solve2)
btn3.grid(row=2, column=2, ipadx=5, ipady=5)

window.mainloop()

# Tạo một tệp Excel mới
workbook = openpyxl.Workbook()

# Chọn sheet mặc định
sheet = workbook.active

# Đặt tiêu đề cho các cột
sheet['A1'] = 'Tổng số SV'
sheet['B1'] = 'Số SV A+'
sheet['C1'] = 'Số sinh viên F'
sheet['D1'] = 'Số điểm mà sinh viên đạt được nhiều nhất'

# Điền dữ liệu vào các ô
data = [
    (60, 32, 10, 9),
    (60, 30, 11, 8),
    (60, 25, 13, 8),
    (60, 22, 12, 7),
]

for row, row_data in enumerate(data, start=2):
    for col, value in enumerate(row_data, start=1):
        sheet.cell(row=row, column=col, value=value)

# Đặt chiều rộng của các cột sao cho dài bằng tiêu đề của cột
for col in sheet.iter_cols(min_col=1, max_col=4, min_row=1, max_row=1):
    max_length = 0
    column = col[0].column_letter  # Lấy tên cột (A, B, C, D)
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column].width = adjusted_width

# Tạo một biểu đồ cột
chart = BarChart()
chart.title = 'Biểu đồ điểm sinh viên'
chart.x_axis.title = 'Cột dữ liệu'
chart.y_axis.title = 'Giá trị'

# Tạo tham chiếu đến dữ liệu cột và thêm nó vào biểu đồ
data_ref = Reference(sheet, min_col=1, min_row=1, max_col=4, max_row=5)
category_ref = Reference(sheet, min_col=1, min_row=2, max_row=5)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(category_ref)

# Thêm biểu đồ vào tệp Excel và lùi sang bên phải 5 cột
sheet.add_chart(chart, "G5")

# Lưu tệp Excel
workbook.save('danh_sach_sinh_vien.xlsx')

# Đóng tệp Excel
workbook.close()

